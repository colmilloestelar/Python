#Nombre Pais, codigo pais, indicador nombre, indicador codigo, 1960 hasta 2017
#program developed by Raul Villar.
# main focus is extract population data and sort it in the required format.
import openpyxl
#main file
doc = openpyxl.load_workbook('DatosPais.xlsx')

sheet = doc.active

#aux var
varPais = sheet['A']
varCode = sheet['B']
varIN = sheet['C']
varIC = sheet['D']
varPoblacion = sheet['E2:BJ1057']

#stored list
pais = []
codigo = []
ini = []
ic = []
poblacion = []

#Year col 1960 to 2017:

init = 1960
fin = 2017
actual = init
anhos = []
for i in range(1,61249):
    anhos.append(actual)
    actual += 1
    if(actual == 2018):
        actual = 1960
#------------------
#populate lists from original data

for v1 in varPais:
    if (v1.value != None):
        pais += 58* [v1.value]
        #pais.append(v1.value)

for v1 in varCode:
    if (v1.value != None):
        codigo += 58* [v1.value]

for v1 in varIN:
    if (v1.value != None):
        ini += 58* [v1.value]

for v1 in varIC:
    if (v1.value != None):
        ic += 58* [v1.value]

for x in range(1,1057):
    poblacion+=([sheet.cell(row=x,column=i).value for i in range(5,63)])

#DEBUG:

#print('x'+str(poblacion[1]))
#print('z'+str(poblacion[2]))
#print('z'+str(poblacion[58]))
#print('z'+str(poblacion[59]))

# -- Write to file:

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

book = Workbook()

sheet = book.active
sheet.title = "Datos Paises 1960 - 2017"

#Country:
col=1
for i, value in enumerate(pais):
    sheet.cell(column=col, row=i+2, value=value)
#Code
col=2
for i, value in enumerate(codigo):
    sheet.cell(column=col, row=i+2, value=value)
#Ini
col=3
for i, value in enumerate(ini):
    sheet.cell(column=col, row=i+2, value=value)
#IC
col=4
for i, value in enumerate(ic):
    sheet.cell(column=col, row=i+2, value=value)
#años
col=5
for i, value in enumerate(anhos):
    sheet.cell(column=col, row=i+2, value=value)
#Data
col=6
for i, value in enumerate(poblacion):
    sheet.cell(column=col, row=i+2, value=value)


print("DONE")

book.save("out.xlxs")




